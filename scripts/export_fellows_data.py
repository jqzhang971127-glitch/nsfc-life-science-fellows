from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parents[1]
FELLOW_ROOT = PROJECT_ROOT.parents[1]
OUTPUT_DIR = PROJECT_ROOT / "public" / "data"
URL_RE = re.compile(r"https?://[^\s;；]+", re.IGNORECASE)


@dataclass(frozen=True)
class DepartmentSource:
    id: str
    name: str
    short_name: str
    prefix: str
    workbook: Path
    output_name: str
    accent: str


DEPARTMENTS = (
    DepartmentSource(
        id="life",
        name="生命科学部",
        short_name="生命科学",
        prefix="C",
        workbook=PROJECT_ROOT.parent / "生命科学学部_v3.xlsx",
        output_name="life.json",
        accent="green",
    ),
    DepartmentSource(
        id="medical",
        name="医学科学部",
        short_name="医学科学",
        prefix="H",
        workbook=FELLOW_ROOT / "医学部_子代码" / "医学科学部_v2.xlsx",
        output_name="medical.json",
        accent="red",
    ),
    DepartmentSource(
        id="information",
        name="信息科学部",
        short_name="信息科学",
        prefix="F",
        workbook=FELLOW_ROOT / "信息科学学部_子代码" / "信息科学学部v1.xlsx",
        output_name="information.json",
        accent="blue",
    ),
)

REQUIRED_HEADERS = (
    "序号",
    "科学部编号",
    "所属学科",
    "现行一级代码（判断）",
    "研究方向",
    "申请人",
    "依托单位",
    "现行一级分类（判断）",
    "当选杰青年份",
    "判断说明",
    "来源URL",
    "来源（可点击）",
    "核验状态",
    "核验证据URL",
    "核验说明",
    "可信度",
)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def extract_urls(values: Iterable[Any]) -> list[str]:
    urls: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = clean_text(value)
        if not text:
            continue
        for match in URL_RE.findall(text):
            url = match.rstrip(".,，。；;）)]}>\"'")
            key = url.rstrip("/")
            if url and key not in seen:
                seen.add(key)
                urls.append(url)
    return urls


def normalized_category(code: str, value: Any) -> str:
    category = clean_text(value)
    if code and category.upper().startswith(code):
        category = category[len(code) :].lstrip(" ：:-—")
    return category


def workbook_timestamp(path: Path) -> str:
    return datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat()


def build_department(source: DepartmentSource) -> dict[str, Any]:
    if not source.workbook.exists():
        raise FileNotFoundError(f"找不到源工作簿：{source.workbook}")

    workbook = openpyxl.load_workbook(source.workbook, data_only=True, read_only=False)
    sheet = workbook.active
    headers = [clean_text(cell.value) for cell in sheet[1]]
    missing_headers = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing_headers:
        raise ValueError(f"{source.workbook.name} 缺少列：{', '.join(missing_headers)}")

    columns = {header: headers.index(header) + 1 for header in REQUIRED_HEADERS}
    valid_code_re = re.compile(rf"^{source.prefix}\d{{2}}$")
    records: list[dict[str, Any]] = []
    code_name_votes: dict[str, Counter[str]] = defaultdict(Counter)

    for row_number in range(2, sheet.max_row + 1):
        name = clean_text(sheet.cell(row_number, columns["申请人"]).value)
        if not name:
            continue

        def value(header: str) -> Any:
            return sheet.cell(row_number, columns[header]).value

        raw_code = clean_text(value("现行一级代码（判断）")).upper()
        current_code = raw_code if valid_code_re.fullmatch(raw_code) else ""
        raw_status = clean_text(value("核验状态"))
        category = normalized_category(current_code, value("现行一级分类（判断）"))
        if current_code and category:
            code_name_votes[current_code][category] += 1

        if current_code:
            classification_state = "classified"
        elif "非医学部" in raw_status or "非信息" in raw_status or "非生命" in raw_status:
            classification_state = "mismatch"
        else:
            classification_state = "unassigned"

        verification_state = "verified" if raw_status == "已核验" else "pending"
        year_value = value("当选杰青年份")
        try:
            award_year = int(year_value)
        except (TypeError, ValueError) as error:
            raise ValueError(f"{source.workbook.name} 第{row_number}行获批年度无效：{year_value}") from error

        source_link = sheet.cell(row_number, columns["来源（可点击）"])
        source_raw = sheet.cell(row_number, columns["来源URL"])
        evidence_link = sheet.cell(row_number, columns["核验证据URL"])
        source_urls = extract_urls(
            (
                source_raw.value,
                source_link.value,
                source_raw.hyperlink.target if source_raw.hyperlink else "",
                source_link.hyperlink.target if source_link.hyperlink else "",
            )
        )
        evidence_urls = extract_urls(
            (
                evidence_link.value,
                evidence_link.hyperlink.target if evidence_link.hyperlink else "",
            )
        )

        record = {
            "rowNumber": row_number,
            "id": f"{source.id}-{award_year}-{name}-{row_number}",
            "departmentId": source.id,
            "departmentName": source.name,
            "scientificDepartmentNumber": clean_text(value("科学部编号")),
            "historicalDiscipline": clean_text(value("所属学科")),
            "currentCode": current_code,
            "rawCurrentCode": raw_code,
            "researchDirection": clean_text(value("研究方向")),
            "name": name,
            "institution": clean_text(value("依托单位")),
            "currentCategory": category,
            "awardYear": award_year,
            "classificationReason": clean_text(value("判断说明")),
            "classificationState": classification_state,
            "sourceUrls": source_urls,
            "sourceLabel": clean_text(source_link.value) or clean_text(source_raw.value),
            "verificationState": verification_state,
            "rawVerificationStatus": raw_status,
            "evidenceUrls": evidence_urls,
            "verificationNote": clean_text(value("核验说明")),
            "confidence": clean_text(value("可信度")),
        }
        records.append(record)

    records.sort(
        key=lambda item: (
            -item["awardYear"],
            item["currentCode"] or "ZZZ",
            item["name"],
            item["rowNumber"],
        )
    )

    code_names = {
        code: votes.most_common(1)[0][0]
        for code, votes in sorted(code_name_votes.items())
        if votes
    }
    duplicate_groups: dict[tuple[str, int, str], list[int]] = defaultdict(list)
    for record in records:
        duplicate_groups[(record["name"], record["awardYear"], record["institution"])].append(
            record["rowNumber"]
        )
    quality_warnings = [
        {
            "type": "duplicate-name-year-institution",
            "name": key[0],
            "awardYear": key[1],
            "institution": key[2],
            "rowNumbers": rows,
            "message": "同名、同年、同单位存在多条记录，请回查源工作簿。",
        }
        for key, rows in duplicate_groups.items()
        if len(rows) > 1
    ]

    years = [record["awardYear"] for record in records]
    classified_count = sum(record["classificationState"] == "classified" for record in records)
    unassigned_count = sum(record["classificationState"] == "unassigned" for record in records)
    mismatch_count = sum(record["classificationState"] == "mismatch" for record in records)
    verified_count = sum(record["verificationState"] == "verified" for record in records)
    code_keys = sorted(code_names)
    manifest = {
        "departmentId": source.id,
        "departmentName": source.name,
        "shortName": source.short_name,
        "accent": source.accent,
        "sourceFile": source.workbook.name,
        "sourceLastModified": workbook_timestamp(source.workbook),
        "generatedAt": datetime.now(tz=timezone.utc).isoformat(),
        "codeVersion": "2026",
        "codePrefix": source.prefix,
        "codeNames": code_names,
        "codeCount": len(code_names),
        "codeMin": code_keys[0] if code_keys else "",
        "codeMax": code_keys[-1] if code_keys else "",
        "fullRecordCount": len(records),
        "classifiedRecordCount": classified_count,
        "unassignedRecordCount": unassigned_count,
        "mismatchRecordCount": mismatch_count,
        "classificationExceptionCount": unassigned_count + mismatch_count,
        "verifiedRecordCount": verified_count,
        "pendingRecordCount": len(records) - verified_count,
        "missingDirectionCount": sum(not record["researchDirection"] for record in records),
        "yearStart": min(years) if years else None,
        "yearEnd": max(years) if years else None,
        "qualityWarningCount": len(quality_warnings),
        "qualityWarnings": quality_warnings,
    }
    return {"manifest": manifest, "records": records}


def write_json(path: Path, payload: Any) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="从三份杰青Excel生成网站JSON数据")
    parser.add_argument("--output", type=Path, default=OUTPUT_DIR)
    args = parser.parse_args()
    args.output.mkdir(parents=True, exist_ok=True)

    summaries: list[dict[str, Any]] = []
    generated: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for source in DEPARTMENTS:
        payload = build_department(source)
        duplicate_ids = [record["id"] for record in payload["records"] if record["id"] in seen_ids]
        if duplicate_ids:
            raise ValueError(f"跨学部记录ID重复：{duplicate_ids[:5]}")
        seen_ids.update(record["id"] for record in payload["records"])
        write_json(args.output / source.output_name, payload)
        generated.append(payload)
        manifest = payload["manifest"]
        summaries.append(
            {
                key: manifest[key]
                for key in (
                    "departmentId",
                    "departmentName",
                    "shortName",
                    "accent",
                    "sourceFile",
                    "sourceLastModified",
                    "codeVersion",
                    "codePrefix",
                    "codeCount",
                    "codeMin",
                    "codeMax",
                    "fullRecordCount",
                    "classifiedRecordCount",
                    "classificationExceptionCount",
                    "verifiedRecordCount",
                    "pendingRecordCount",
                    "yearStart",
                    "yearEnd",
                    "qualityWarningCount",
                )
            }
            | {"dataFile": source.output_name}
        )

    department_manifest = {
        "generatedAt": datetime.now(tz=timezone.utc).isoformat(),
        "codeVersion": "2026",
        "departmentCount": len(summaries),
        "totalRecordCount": sum(item["manifest"]["fullRecordCount"] for item in generated),
        "totalClassifiedRecordCount": sum(
            item["manifest"]["classifiedRecordCount"] for item in generated
        ),
        "totalClassificationExceptionCount": sum(
            item["manifest"]["classificationExceptionCount"] for item in generated
        ),
        "departments": summaries,
    }
    write_json(args.output / "departments.json", department_manifest)

    print(
        "生成完成："
        + "，".join(
            f"{item['departmentName']}{item['fullRecordCount']}条"
            for item in summaries
        )
    )


if __name__ == "__main__":
    main()
